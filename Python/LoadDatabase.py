'''
ESTA FUNCIÓN DEBE RECIBIR POR PARÁMETROS EL FICHERO EXCEL A CARGAR EN LA BASE DE DATOS, ADEMÁS DE LAS COLUMNAS Y FILAS DE REFERENCIA 
LA FUNCIÓN DEBE RECORRER DICHO EXCEL E INSERTAR EN BASE DE DATOS TODAS LAS CLÁUSULAS COMPROBANDO SI LA CLÁUSULA YA ESTÁ EN BBDD

#Ejemplo: python LoadDatabase.py 'C:/Users/mcuberos/Desktop/AppGestorRequisitos_old/Python/D0000016800_EEFAE SALOON HVAC P2_Ed-HC_230517.xlsx' '7' 'G'''

import os
import sqlite3
import re
import pandas as pd
from tkinter import messagebox
import sys
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askopenfilename 
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import pyodbc 
import json
from datetime import date



def leer_configuracion(fichero_config):
    with open(fichero_config, 'r') as archivo:
        configuracion = json.load(archivo)
    return configuracion

#PathName=os.path.dirname(__file__)
fileConfig="C:/AppGestorRequisitos/config.json"

configuracion=leer_configuracion(fileConfig)

# Acceder a los datos
direccion_bd = configuracion['database']['host']
servidor=direccion_bd.replace("/","\\")
database=configuracion['database']['database_name']
username=configuracion['database']['username']
tablename=configuracion['database']['table_name']
password='Prueb@sManuel23'
# Cadena de conexión
connection_string = f'DRIVER={{SQL Server}};SERVER={servidor};DATABASE={database};UID={username};PWD={password}'
connection = pyodbc.connect(connection_string)
cursor = connection.cursor()
cursor.execute("SELECT * FROM T_REQUISITOS")
listadoRequisitos=cursor.fetchall()

fecha_actual_sql = date.today().strftime('%Y-%m-%d')

dftemp = pd.DataFrame({'Id_Req': [],
                   'Desc_Req': [],
                   'Nueva_Respuesta':[],
                   'Nuevos_Comentarios':[],
                   'Nuevo_Proyecto':[],
                   'Nuevo_Fichero':[],
                   'Tipo_Vehiculo':[],
                   'Entregable_CBC':[],
                   '%Coincidencia':[],
                   'ID_Req_BBDD': [],
                   'Descr_Req_BBDD':[],
                   'Resp_BBDD':[],
                   'Coment_BBDD':[],
                   'Proyecto_Origen':[],
                   'Fichero_Origen':[],
                   'Vehiculo_BBDD':[],
                   'Entregable_CBC_BBDD':[]})

def InsertClause(clausula):
    """InsertClause INTENTA INSERTAR UNA CLÁUSULA EN BBDD.
    ANTES DE INSERTARLA DEBE COMBROBAR SI LA CLÁUSULA YA EXISTE LLAMANDO A LA FUNCIÓN CHECKCLAUSE"""
    global dftemp
    global tipo_vehiculo
    global fichero_origen
    global proy_origen
    global entregable_cbc
    global connection_string
    global fecha_actual_sql
    global listadoRequisitos
    
    '''
    # Si quiero comprobar si en el mismo fichero que quiero subir a bbdd hay clausulas repetidas, establezco una nueva conexión a bbdd para cada clausula
    connection = pyodbc.connect(connection_string)

    cursor = connection.cursor()

    cursor.execute("SELECT * FROM T_REQUISITOS")
    listadoRequisitos=cursor.fetchall()
    '''
    ExisteClausula=FALSE

    for requisito in listadoRequisitos:
        if clausula[1]==requisito[1]:
            if clausula[3]==requisito[3] and clausula[4]==requisito[4]:
                print("La cláusula situada en la fila", clausula[0]+1  ,"con identificador ",clausula[1]," ya existe en bbdd con el ID_REQ:",requisito[1])
            else:
                nueva_fila = pd.Series([clausula[1], clausula[2],clausula[3],clausula[4],proy_origen,fichero_origen,tipo_vehiculo,entregable_cbc,100,requisito[1],requisito[2],requisito[3],requisito[4],requisito[10],requisito[11],requisito[5],requisito[12]], index=dftemp.columns)
                dftemp = dftemp._append(nueva_fila, ignore_index=True)
        else:
            accuracy=CheckClause(clausula[2],requisito[2]) #La salida de la función debe ser un % de coincidencia (accuracy) entre los dos requisitos. if accuracy<60%, insert requisito en bbdd
            if accuracy>85:
                ExisteClausula=TRUE
                print("La cláusula situada en la fila", clausula[0]+1  ,"con identificador ",clausula[4]," ya existe en bbdd con el id ",requisito[0],"-ID_REQ:",requisito[1], "con un porcentaje de coincidencia del ",accuracy," por ciento")
                nueva_fila = pd.Series([clausula[1], clausula[2],clausula[3],clausula[4],proy_origen,fichero_origen,tipo_vehiculo,entregable_cbc,accuracy,requisito[1],requisito[2],requisito[3],requisito[4],requisito[10],requisito[11],requisito[5],requisito[12]], index=dftemp.columns)
                dftemp = dftemp._append(nueva_fila, ignore_index=True)


    if ExisteClausula==FALSE:
        try:
            cursor.execute("INSERT INTO T_REQUISITOS VALUES (?, ?, ?, ?,?,'CAF',?,NULL,1,?,?,?)",(clausula[1],clausula[2],clausula[3],clausula[4],tipo_vehiculo,fecha_actual_sql,proy_origen,fichero_origen,entregable_cbc))
            connection.commit()
        except Exception as e:
            messagebox.showinfo("¡ATENCIÓN!","NO SE HA PODIDO CREAR EL REGISTRO EN BASE DE DATOS: " + str(e))

    return dftemp


def CheckClause(newClause,requirement):
    """ESTA FUNCIÓN COMPRUEBA SI UN REQUISITO NUEVO ES IGUAL A OTRO GUARDADO EN BBDD.
        DEVUELVE UN % DE COINDICENCIA ENTRE LOS DOS REQUISITOS"""
    if newClause==requirement:
        accuracy=100
    else:
        accuracy=0
        longParcial=30
        numTramos=int(len(newClause)/longParcial)+1
        #longParcial=int(len(newClause)/15)
        for aux in range(numTramos):
            cadena=newClause[aux*longParcial:longParcial*(aux+1)]
            if re.search(re.escape(cadena),requirement):
                accuracy=accuracy+99/numTramos
                    
    return accuracy

print("SELECCIONE EL CBC QUE DESEA CARGAR A BASE DE DATOS")
fileName=askopenfilename()
filaHeader=input("INDIQUE LA FILA DONDE SE ENCUENTRA LA CABECERA DEL CBC ")
colIdReq=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LOS IDs DEL REQUISITO (A,B,C,D,...) ")
colClause=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LAS DESCRIPCIONES DE LOS REQUISITOS (A,B,C,D,...) ") #para convertir la columna (letra) a número: ord(colClause.lower())-96
colResp=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LAS RESPUESTAS A LOS REQUISITOS - C/NC/NA (A,B,C,D,...) ")
colComments=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LOS COMENTARIOS (A,B,C,D,...) ")
nombre_hoja=input("INDIQUE EL NOMBRE DE LA HOJA DONDE SE ENCUENTRAN LOS REQUISITOS (POR DEFECTO, Requirements) ")
if nombre_hoja=="":
    nombre_hoja="Requirements"
tipo_vehiculo=input("INDIQUE EL TIPO DE VEHÍCULO DEL CBC (TRANVÍA, METRO, ...) ")
proy_origen=input("INDIQUE EL NOMBRE DEL PROYECTO: ")
fichero_origen=os.path.split(fileName)[1]
entregable_cbc=input("INDIQUE EL TIPO DE EQUIPO AL QUE HACE REFERENCIA EL CBC (SALA, CABINA, GENERAL): ")

df=pd.read_excel(fileName, sheet_name=nombre_hoja,header=int(filaHeader)-1,keep_default_na=FALSE)
#EL DATAFRAME CONSIDERA LOS "NA" COMO NaN, ASÍ QUE TRATO LOS NA DE LA COLUMNA RESPUESTA PARA QUE LOS GUARDE CORRECTAMENTE CON LA FUNCIÓN fillna
#TituloColumnaRespuesta=df.columns[(ord(colResp.lower())-97)]
#df[TituloColumnaRespuesta]=df[TituloColumnaRespuesta].fillna("NA")
#TituloColumnaClausula=df.columns[(ord(colClause.lower())-97)]
#df[TituloColumnaClausula]=df[TituloColumnaClausula].fillna("NA")


#RECORRO EL DATAFRAME EN LA COLUMNA colClause, OMITIENDO LAS FILAS QUE SE CORRESPONDEN A TÍTULOS 
for kk in range(len(df)):
    if (df.iloc[kk][(ord(colResp.lower())-97)]!="" and len(df.iloc[kk][(ord(colClause.lower())-97)])>5):
        #defino la variable clausula como una tupla que contiene la descripción de la clausula, la respuesta, comentarios y req. familia
        clausula=(kk,df.iloc[kk][(ord(colIdReq.lower())-97)],df.iloc[kk][(ord(colClause.lower())-97)],df.iloc[kk][(ord(colResp.lower())-97)],df.iloc[kk][(ord(colComments.lower())-97)])
        InsertClause(clausula)


if len(dftemp)>0:
    Ruta=os.path.dirname(fileName) #busco la ruta donde se debe encontrar la bbdd
    excelTemp=(Ruta + "/ExcelTemp.xlsx")
    
    writer = ExcelWriter(excelTemp)
    dftemp.to_excel(writer, 'Requirements Temp', index=False)
    writer.close()
    #DOY FORMATO A COLUMNAS 
    libro=load_workbook(excelTemp)
    hoja=libro.active
    print("Dando formato a excel temp...")
    columnas_a_formatear = ["A","B","C","D","E","F","G","H"]
    fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for col in columnas_a_formatear:
        for celda in hoja[col]:
            celda.fill=fill
    
    fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    for celda in hoja["I"]:
        celda.fill=fill

    columnas_a_formatear = ["J","K","L","M","N","O","P","Q"]
    fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    for col in columnas_a_formatear:
        for celda in hoja[col]:
            celda.fill=fill


    filas_borrar=[]
    #resalto las cláusulas que aparezcan repetidas
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for i in range(len(hoja["A"])):
        if hoja["A"+str(i+1)].value==hoja["A"+str(i+2)].value:
            if hoja["G"+str(i+1)].value==hoja["P"+str(i+1)].value and  hoja["H"+str(i+1)].value==hoja["Q"+str(i+1)].value: #si en la primera fila coincide tipo veh y entregable, borro la otra fila
                filas_borrar.append(i+2)
            elif hoja["G"+str(i+2)].value==hoja["P"+str(i+2)].value and  hoja["H"+str(i+2)].value==hoja["Q"+str(i+2)].value: #y viceversa
                filas_borrar.append(i+1)
            else:
                hoja["A"+str(i+1)].fill=fill
                hoja["A"+str(i+2)].fill=fill

        if hoja["C"+str(i+1)].value!=hoja["L"+str(i+1)].value:
            hoja["C"+str(i+1)].fill=fill
            hoja["L"+str(i+1)].fill=fill

        if hoja["D"+str(i+1)].value!=hoja["M"+str(i+1)].value:
            hoja["D"+str(i+1)].fill=fill
            hoja["M"+str(i+1)].fill=fill

    fillHeader = PatternFill(start_color="EF9191", end_color="EF9191", fill_type="solid")
    font = Font(bold=True)
    encabezado=hoja[1]
    for celda in encabezado:
        celda.fill=fillHeader
        celda.font=font

    for f in filas_borrar[::-1]:
        hoja.delete_rows(f)

    libro.save(excelTemp)

    messagebox.showinfo("EXCEL TEMPORAL CREADO","SE HA CREADO UN EXCEL TEMPORAL CON LAS CLÁUSULAS CON COINCIDENCIAS EN LA RUTA " + Ruta)   


messagebox.showinfo("FIN PROCESO","EJECUCIÓN FINALIZADA CON ÉXITO")