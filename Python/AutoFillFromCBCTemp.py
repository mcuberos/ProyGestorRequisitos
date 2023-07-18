######## FUNCIÓN PARA CARGAR UN CBC Y AUTORELLENAR LAS CLÁUSULAS EN BASE A LO QUE HAYA INFORMADO EN EL CBC TEMPORAL QUE SE HAYA GENERADO PREVIAMENTE


import os
import sqlite3
import re
import pandas as pd
from tkinter import messagebox
import sys
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askopenfilename 
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font



print("SELECCIONE EL FICHERO CBC A RELLENAR")
fileCBC=askopenfilename()
filaHeader=input("INDIQUE LA FILA DONDE SE ENCUENTRA LA CABECERA DEL CBC ")
colIdReq=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LOS IDs DEL REQUISITO (A,B,C,D,...) ")
colClause=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LAS DESCRIPCIONES DE LOS REQUISITOS (A,B,C,D,...) ") #para convertir la columna (letra) a número: ord(colClause.lower())-96
colResp=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LAS RESPUESTAS A LOS REQUISITOS - C/NC/NA (A,B,C,D,...) ")
colComments=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LOS COMENTARIOS (A,B,C,D,...) ")
nombre_hoja=input("INDIQUE EL NOMBRE DE LA HOJA DONDE SE ENCUENTRAN LOS REQUISITOS (POR DEFECTO, Requirements)")
if nombre_hoja=="":
    nombre_hoja="Requirements"



print("SELECCIONE EL FICHERO TEMPORAL GENERADO PREVIAMENTE")
fileTemp=askopenfilename()



df=pd.read_excel(fileTemp, sheet_name="CBC Temporal",keep_default_na=FALSE)

#COMPRUEBO SI TENGO CLÁUSULAS REPETIDAS EN EL DATAFRAME TEMPORAL
duplicates=df['Id_Req'].duplicated()
duplicated_values=df['Id_Req'][duplicates].unique()
if duplicates.any():
    print("HAY CLÁUSULAS REPETIDAS EN EL EXCEL TEMPORAL. SE DEBE CORREGIR PARA PODER EJECUTAR EL PROGRAMA")
    for kk in range(len(duplicated_values)):
        print("LA CLÁUSULA "+ duplicated_values[kk] +" SE ENCUENTRA DUPLICADA." )
        input()
    sys.exit()
    

libro=load_workbook(filename=fileCBC)
hoja=libro[nombre_hoja]
df_cbc=pd.read_excel(fileCBC, sheet_name=nombre_hoja,header=int(filaHeader)-1,keep_default_na=FALSE)

#RECORRO EL DATAFRAME DEL EXCEL TEMPORAL. SI ENCUENTRO UNA CLÁUSULA REPETIDA HAY QUE ALERTAR AL USUARIO 
for kk in range(len(df)):
        #defino la variable clausula como una tupla que contiene: id_requisito, descripción de la clausula, la respuesta, comentarios
        clausula=(df.iloc[kk][0],df.iloc[kk][1],df.iloc[kk][7],df.iloc[kk][8])

        for fila in range(len(df_cbc)):
            celdaReq=hoja[colIdReq + str(int(filaHeader)+fila+1)]
            celdaResp=hoja[colResp + str(int(filaHeader)+fila+1)]
            celdaComm=hoja[colComments + str(int(filaHeader)+fila+1)]
            if celdaReq.value==clausula[0]:
                celdaResp.value=clausula[2]
                celdaComm.value=clausula[3]   
        
libro.save(fileCBC)       


messagebox.showinfo("FIN PROCESO","EJECUCIÓN FINALIZADA CON ÉXITO")